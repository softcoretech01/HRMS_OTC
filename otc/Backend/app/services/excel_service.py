from typing import List, Optional
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from sqlalchemy.orm import Session
from app.models.employee_model import Employee
from app.database import SessionLocal
from io import BytesIO
import pandas as pd
import os

EXCEL_FILE_PATH = "employee_data.xlsx"  # single workbook for all exports (employees + attendance)
HEADERS = ["emp_id", "name", "designation", "salary", "department", "hod", "supervisor", "status", "created_at", "updated_at"]

# Attendance specific headers (written to a separate sheet in the same workbook)
# Includes computed fields: working_hours, ot_hours, ot (Yes/No)
ATTENDANCE_HEADERS = [
    "emp_id",
    "name",
    "clock_in",
    "clock_out",
    "working_hours",
    "ot_hours",
    "ot",
    "created_at",
    "updated_at",
]


def _get_or_create_sheet(wb: Workbook, sheet_name: str, headers: List[str]):
    """Return a worksheet with given name; create and populate headers if missing."""
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Ensure headers are present and updated to the provided headers (uppercased)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header.upper()
            cell.font = Font(bold=True)
        return ws

    ws = wb.create_sheet(title=sheet_name)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header.upper()
        cell.font = Font(bold=True)
    return ws

def create_or_append_to_excel(employees: List[Employee]) -> str:
    """Create a new Excel file or append to existing one with employee data.
    Returns the path to the Excel file."""
    sheet_name = "Employee Data"
    # If file exists, load it and collect existing emp_id values from the employee sheet
    if os.path.exists(EXCEL_FILE_PATH):
        wb = load_workbook(EXCEL_FILE_PATH)
        ws = _get_or_create_sheet(wb, sheet_name, HEADERS)

        # Collect existing emp_id values from column 1
        existing_ids = set()
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row and row[0] is not None:
                existing_ids.add(str(row[0]))

        # Filter employees to only include ones not already present
        new_employees = [e for e in employees if e.emp_id and str(e.emp_id) not in existing_ids]

        # If no new employees to append, return existing file path
        if not new_employees:
            return EXCEL_FILE_PATH

        start_row = ws.max_row + 1
    else:
        # Create new workbook and employee sheet with headers
        wb = Workbook()
        ws = _get_or_create_sheet(wb, sheet_name, HEADERS)
        new_employees = employees
        start_row = 2

    # Add new data rows only
    for offset, emp in enumerate(new_employees):
        row = start_row + offset
        ws.cell(row=row, column=1, value=emp.emp_id)
        ws.cell(row=row, column=2, value=emp.name)
        ws.cell(row=row, column=3, value=emp.designation)
        ws.cell(row=row, column=4, value=emp.salary)
        ws.cell(row=row, column=5, value=emp.department)
        ws.cell(row=row, column=6, value=emp.hod)
        ws.cell(row=row, column=7, value=emp.supervisor)
        ws.cell(row=row, column=8, value=getattr(emp, "status", None))
        # created_at may be None or a datetime; guard accordingly
        if getattr(emp, "created_at", None):
            try:
                created_val = emp.created_at.strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                created_val = str(emp.created_at)
        else:
            created_val = None
        ws.cell(row=row, column=9, value=created_val)
        # updated_at may be None or a datetime; guard accordingly
        if getattr(emp, "updated_at", None):
            try:
                updated_val = emp.updated_at.strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                updated_val = str(emp.updated_at)
        else:
            updated_val = None
        ws.cell(row=row, column=10, value=updated_val)

    # Auto-adjust column widths based on content
    for col_idx in range(1, len(HEADERS) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[col_letter]:
            if cell.value is not None:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        # Add a small padding and limit max width
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # Save the workbook (single master file)
    target_path = os.path.abspath(EXCEL_FILE_PATH)
    try:
        wb.save(target_path)
        return target_path
    except PermissionError:
        # Common on Windows when the file is open in Excel. Fall back to a timestamped file
        ts = datetime.now().strftime("%Y%m%d%H%M%S")
        fallback_name = f"employee_data_{ts}.xlsx"
        fallback_path = os.path.abspath(fallback_name)
        try:
            wb.save(fallback_path)
            return fallback_path
        except Exception as exc:
            # If even the fallback fails, raise so caller can return a 500 with context
            raise


def write_employees_to_excel(employees: List[Employee]) -> str:
    """Overwrite the Excel file with the provided employees list.
    This is intended to be used after updates/deletes so the file exactly matches DB state.
    Returns the absolute path to the saved file."""
    sheet_name = "Employee Data"
    # Load existing workbook to preserve other sheets (attendance). If none, create new.
    if os.path.exists(EXCEL_FILE_PATH):
        wb = load_workbook(EXCEL_FILE_PATH)
        # Remove existing employee sheet if present to overwrite
        if sheet_name in wb.sheetnames:
            std = wb[sheet_name]
            wb.remove(std)
        ws = _get_or_create_sheet(wb, sheet_name, HEADERS)
    else:
        wb = Workbook()
        ws = _get_or_create_sheet(wb, sheet_name, HEADERS)

    # Write each employee as a full row
    for idx, emp in enumerate(employees, start=2):
        ws.cell(row=idx, column=1, value=emp.emp_id)
        ws.cell(row=idx, column=2, value=emp.name)
        ws.cell(row=idx, column=3, value=emp.designation)
        ws.cell(row=idx, column=4, value=emp.salary)
        ws.cell(row=idx, column=5, value=emp.department)
        ws.cell(row=idx, column=6, value=emp.hod)
        ws.cell(row=idx, column=7, value=emp.supervisor)
        ws.cell(row=idx, column=8, value=getattr(emp, "status", None))
        if getattr(emp, "created_at", None):
            try:
                created_val = emp.created_at.strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                created_val = str(emp.created_at)
        else:
            created_val = None
        ws.cell(row=idx, column=9, value=created_val)
        if getattr(emp, "updated_at", None):
            try:
                updated_val = emp.updated_at.strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                updated_val = str(emp.updated_at)
        else:
            updated_val = None
        ws.cell(row=idx, column=10, value=updated_val)

    # Auto-adjust column widths
    for col_idx in range(1, len(HEADERS) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[col_letter]:
            if cell.value is not None:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    target_path = os.path.abspath(EXCEL_FILE_PATH)
    try:
        wb.save(target_path)
        return target_path
    except PermissionError:
        ts = datetime.now().strftime("%Y%m%d%H%M%S")
        fallback_name = f"employee_data_{ts}.xlsx"
        fallback_path = os.path.abspath(fallback_name)
        wb.save(fallback_path)
        return fallback_path


def create_or_append_attendance_excel(att_records: List[Employee]) -> str:
    """Deprecated append behavior replaced: always write full Attendance sheet to ensure all columns and rows."""
    return write_attendance_to_excel(att_records)


def write_attendance_to_excel(att_records: List[Employee]) -> str:
    """Overwrite the Attendance sheet in the master workbook with provided records
    while preserving the Employee Data sheet."""
    sheet_name = "Attendance"
    # Load or create workbook, preserving Employee Data sheet
    if os.path.exists(EXCEL_FILE_PATH):
        wb = load_workbook(EXCEL_FILE_PATH)
        if "Employee Data" not in wb.sheetnames and wb.sheetnames:
            # If there's a default sheet but no Employee Data, rename it
            wb.active.title = "Employee Data"
    else:
        wb = Workbook()
        if wb.sheetnames:  # Rename default sheet to Employee Data
            wb.active.title = "Employee Data"
    
    # Remove and recreate Attendance sheet
    if sheet_name in wb.sheetnames:
        old = wb[sheet_name]
        wb.remove(old)
    ws = _get_or_create_sheet(wb, sheet_name, ATTENDANCE_HEADERS)

    for idx, rec in enumerate(att_records, start=2):
        ws.cell(row=idx, column=1, value=getattr(rec, "emp_id", None))
        ws.cell(row=idx, column=2, value=getattr(rec, "name", None))
        ci = getattr(rec, "clock_in", None)
        co = getattr(rec, "clock_out", None)
        try:
            ws.cell(row=idx, column=3, value=ci.strftime("%Y-%m-%d %H:%M:%S") if ci else None)
        except Exception:
            ws.cell(row=idx, column=3, value=str(ci) if ci else None)
        try:
            ws.cell(row=idx, column=4, value=co.strftime("%Y-%m-%d %H:%M:%S") if co else None)
        except Exception:
            ws.cell(row=idx, column=4, value=str(co) if co else None)
        # computed columns
        ws.cell(row=idx, column=5, value=getattr(rec, "working_hours", None))
        ws.cell(row=idx, column=6, value=getattr(rec, "ot_hours", None))
        ot_val = getattr(rec, "ot", None)
        ws.cell(row=idx, column=7, value=("Yes" if ot_val else ("No" if ot_val is not None else None)))
        # timestamps
        ca = getattr(rec, "created_at", None)
        try:
            ws.cell(row=idx, column=8, value=ca.strftime("%Y-%m-%d %H:%M:%S") if ca else None)
        except Exception:
            ws.cell(row=idx, column=8, value=str(ca) if ca else None)
        ua = getattr(rec, "updated_at", None)
        try:
            ws.cell(row=idx, column=9, value=ua.strftime("%Y-%m-%d %H:%M:%S") if ua else None)
        except Exception:
            ws.cell(row=idx, column=9, value=str(ua) if ua else None)

    for col_idx in range(1, len(ATTENDANCE_HEADERS) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[col_letter]:
            if cell.value is not None:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    target_path = os.path.abspath(EXCEL_FILE_PATH)
    try:
        wb.save(target_path)
        return target_path
    except PermissionError:
        ts = datetime.now().strftime("%Y%m%d%H%M%S")
        fallback_name = f"employee_data_{ts}.xlsx"
        fallback_path = os.path.abspath(fallback_name)
        wb.save(fallback_path)
        return fallback_path


def build_attendance_workbook_bytes(att_records) -> bytes:
    """Build an in-memory workbook containing only the Attendance sheet with up-to-date headers and rows.
    Returns the XLSX bytes."""
    wb = Workbook()
    # Remove default sheet if present; create Attendance with headers
    if wb.sheetnames:
        default = wb.active
        wb.remove(default)
    ws = _get_or_create_sheet(wb, "Attendance", ATTENDANCE_HEADERS)

    for idx, rec in enumerate(att_records, start=2):
        ws.cell(row=idx, column=1, value=getattr(rec, "emp_id", None))
        ws.cell(row=idx, column=2, value=getattr(rec, "name", None))
        ci = getattr(rec, "clock_in", None)
        co = getattr(rec, "clock_out", None)
        try:
            ws.cell(row=idx, column=3, value=ci.strftime("%Y-%m-%d %H:%M:%S") if ci else None)
        except Exception:
            ws.cell(row=idx, column=3, value=str(ci) if ci else None)
        try:
            ws.cell(row=idx, column=4, value=co.strftime("%Y-%m-%d %H:%M:%S") if co else None)
        except Exception:
            ws.cell(row=idx, column=4, value=str(co) if co else None)
        ws.cell(row=idx, column=5, value=getattr(rec, "working_hours", None))
        ws.cell(row=idx, column=6, value=getattr(rec, "ot_hours", None))
        ot_val = getattr(rec, "ot", None)
        ws.cell(row=idx, column=7, value=("Yes" if ot_val else ("No" if ot_val is not None else None)))
        ca = getattr(rec, "created_at", None)
        try:
            ws.cell(row=idx, column=8, value=ca.strftime("%Y-%m-%d %H:%M:%S") if ca else None)
        except Exception:
            ws.cell(row=idx, column=8, value=str(ca) if ca else None)
        ua = getattr(rec, "updated_at", None)
        try:
            ws.cell(row=idx, column=9, value=ua.strftime("%Y-%m-%d %H:%M:%S") if ua else None)
        except Exception:
            ws.cell(row=idx, column=9, value=str(ua) if ua else None)

    for col_idx in range(1, len(ATTENDANCE_HEADERS) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[col_letter]:
            if cell.value is not None:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

def get_employee_by_id(db: Session, employee_id: int) -> Optional[Employee]:
    return db.query(Employee).filter(Employee.id == employee_id).first()


def update_employee(db: Session, employee_id: int, data: dict) -> Optional[Employee]:
    """Update an employee by primary `id`. Also regenerates the Excel file to reflect changes.
    Returns the updated Employee or None if not found."""
    emp = get_employee_by_id(db, employee_id)
    if not emp:
        return None

    allowed = {"name", "designation", "salary", "department", "hod", "supervisor", "status"}
    for key, value in data.items():
        if key in allowed:
            setattr(emp, key, value)

    db.add(emp)
    db.commit()
    db.refresh(emp)

    # Regenerate the Excel file to reflect the updated DB
    all_emps = db.query(Employee).all()
    try:
        write_employees_to_excel(all_emps)
    except Exception:
        # Best-effort: do not fail DB update if Excel write fails
        pass

    return emp


def delete_employee(db: Session, employee_id: int) -> bool:
    """Delete an employee by primary `id`. Regenerates Excel after deletion."""
    emp = get_employee_by_id(db, employee_id)
    if not emp:
        return False

    db.delete(emp)
    db.commit()

    # Regenerate full Excel after delete
    all_emps = db.query(Employee).all()
    try:
        write_employees_to_excel(all_emps)
    except Exception:
        pass

    return True


EXCEL_PATH = "employee_data.xlsx"

def create_or_append_to_excel(employees):
    return write_employees_to_excel(employees)


def write_employees_to_excel(employees):
    data = []

    for emp in employees:
        data.append({
            "ID": emp.id,
            "Emp ID": emp.emp_id,
            "Name": emp.name,
            "Designation": emp.designation,
            "Salary": emp.salary,
            "Department": emp.department,
            "HOD": emp.hod,
            "Supervisor": emp.supervisor,
            "Status": emp.status
        })

    df = pd.DataFrame(data)

    if df.empty:
        print("⚠ No employees found to export")

    df.to_excel(EXCEL_PATH, index=False)

    return EXCEL_PATH